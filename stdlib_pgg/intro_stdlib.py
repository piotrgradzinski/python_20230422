"""
Pickle

Is a build in format in python for data storage (serialization and deserialization).
https://docs.python.org/3/library/pickle.html?highlight=pickle#module-pickle
"""

import pickle

data = {
    'year': 2020,
    'students': [
        {
            'first_name': 'John',
            'last_name': 'Doe'
        },
        {
            'first_name': 'Ann',
            'last_name': 'Kowalski'
        }
    ]
}

pickle.dump(data, open('data.p', 'wb'))

"""
Data format are (most of the time) language independent.
!!!
Java != JavaScript
!!!


JSON - JavaScript Object Notation
Data format is derived from JavaScript.
Important to know that it's a separate data format. 
It may look similar to Python structures (like dictionaries or lists)
but it's not the same and has some limitations. 

There are some signification differences between Python structures and JSON structures:
Python - JSON
dict - object
list - array
True/False - true/false
None - null
tuple() - there is no tuple structure in JSON
set() - there is no set structure in JSON

https://en.wikipedia.org/wiki/JSON
"""

data = {
    'year': 2020,
    'students': [
        {
            'first_name': 'John',
            'last_name': 'Doe'
        },
        {
            'first_name': 'Ann',
            'last_name': 'Kowalski'
        }
    ],
    'available': True,
    'is_connected': None,
}

import json

"""
Within json module we have:
- dumps - will return a string in a JSON format
- dump - will write the data in a JSON to a file
- loads - converts JSON data into Python data from a string
- load - converts JSON data into Python data from a file
"""

print(json.dumps(data, indent=4))  # indent - makes indentations, so JSON looks better and is more readable

json.dump(data, open('data.json', 'w'), indent=4)

data = json.loads('{"course": "Python"}')
print(data)

data = json.load(open('employees.json'))
print(data)

print('-' * 30)

"""
CSV - Comma-separated_values
https://en.wikipedia.org/wiki/Comma-separated_values
https://realpython.com/python-csv/
"""

import csv
from pprint import pprint

# reading
with open('data.csv', encoding='utf-8') as file_handle:
    data = csv.DictReader(file_handle, delimiter=';')
    pprint(list(data))

# writing
with open('data2.csv', 'w', encoding='utf-8') as file_handle:
    fields = ['first_name', 'last_name']
    writer = csv.DictWriter(file_handle, fieldnames=fields, delimiter=';')

    writer.writeheader()
    writer.writerow({'first_name': 'Piotr', 'last_name': 'GG'})

print('-' * 30)

"""
Excel files
https://pypi.org/project/openpyxl/
https://openpyxl.readthedocs.io/en/stable/

We have to install openpyxl library, in terminal 
pip install openpyxl

https://realpython.com/openpyxl-excel-spreadsheets-python/
"""
from openpyxl import load_workbook

workbook = load_workbook('employees.xlsx')

# select proper sheet
worksheet = workbook['Arkusz1']

salaries = []
for row in worksheet.iter_rows(min_row=2):
    # each row is a tuple of Cell objects
    print(row[0].value, row[1].value, row[2].value)
    salaries.append(row[2].value)

print(salaries)

salaries_sum = sum(salaries)
salaries_average = salaries_sum / len(salaries)

print(f"Salaries sum: {salaries_sum}")
print(f"Salaries average: {salaries_average}")

print('-' * 30)

"""
API - Application Programming Interface

https://realpython.com/api-integration-in-python/

How we can communicate with other services / applications
using HTTP protocol (pipe) and RestAPI (how we prepare the message).

HTTP protocol
HTTPS - a secure version of a HTTP protocol which uses asymmetric encryption

URL
https://alx.training/en/bootcamp-python-english/
https - protocol
alx.training - domain name
/en/bootcamp-python-english/ - path

To send and receive HTTP request / response we can use:
- urllib - embedded in python
- requests - pip install requests
"""

# urllib
import urllib.request
import json
from pprint import pprint

with urllib.request.urlopen('https://data.police.uk/api/forces') as response:
    data = response.read()  # we are receiving a string
    data = json.loads(data)
    pprint(data)

    for force in data:
        print(force['name'])


"""
requests library
https://pypi.org/project/requests/
pip install requests
"""